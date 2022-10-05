
from ast import Str
from multiprocessing import context
from multiprocessing.sharedctypes import Value
from optparse import Values
from django.http import HttpResponse 
from django.shortcuts import render,redirect,get_object_or_404
from .forms import AddListForm,DeptCreateForm, PlaceCreateForm,DeptEditForm,PlaceEditForm
from .models import DeptMaster, PurchaserMaster,ListOfPurchases
from django.db.models import Q
from functools import reduce
from operator import and_
import openpyxl
from openpyxl.styles import Alignment
from django.contrib import messages
from datetime import datetime
import datetime as d
#from dateutil.relativedelta import relativedelta
from django.db import connection


def index(request):
    with connection.cursor() as cursor:
        cursor.execute("""SELECT item_name,dept,
            (DATEDIFF(day, MIN(order_day), MAX(order_day))/(COUNT(item_name)-1))as order_qycle,
            (DATEADD(day,(DATEDIFF(day, MIN(order_day), MAX(order_day))/(COUNT(item_name)-1)),Max(order_day))) as next_order
            FROM buy_listofpurchases
            LEFT JOIN buy_deptmaster
            ON buy_listofpurchases.dept_id = buy_deptmaster.dept_id
            GROUP BY item_name,dept
            HAVING COUNT(item_name) > = 2 AND  (DATEDIFF(day, MIN(Order_Day), MAX(Order_day))/COUNT(item_name)) >= 1""")
        list = cursor.fetchall()

    now = d.datetime.now()
    year = now.year
    month = now.month
    now_month = []
    before_month = []
    next_month = []

    for order_day in list:
        if order_day[3].month == month and order_day[3].year == year:
            now_month.append(order_day)
        elif order_day[3].month == month+1 and order_day[3].year == year:
            next_month.append(order_day)
        elif order_day[3].month == month-1 and order_day[3].year == year:
            before_month.append(order_day)
    
    print(next_month)

    context ={
        'now':now_month,
        'next':next_month,
        'before':before_month,
    }
    return render(request,'buy/index.html',context)

def dept(request):
    context ={
        'dept_list':DeptMaster.objects.all(),
    }
    return render(request,'buy/dept.html',context)

def add(request,):
    form = AddListForm(request.POST or None) 
    if request.method == 'POST' and form.is_valid():
            form.save()
            messages.add_message(request, messages.SUCCESS, "登録が完了しました。")
            return redirect('buy:add')
    context= {
        
        'item':ListOfPurchases.objects.distinct().values('item_name'),
        'form':form,
    }
    return render(request,'buy/add.html',context)
    

def deptup(request, pk):
    #urlのpkを基にdeptを取得
    dept = get_object_or_404(DeptMaster,pk=pk)
    #フォームに取得したdeptを紐づけ
    form = DeptEditForm(request.POST or None, instance=dept)
    #methot=POST、つまり送信ボタン押下時、入力内容が問題なければ
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.add_message(request, messages.SUCCESS, "更新が完了しました。")
        return redirect('buy:dept')

    context = {
        'form':form,
    }
    return render(request,'buy/dept_form.html',context)


def deptadd(request):
    form = DeptCreateForm(request.POST or None)
    func = 'deptadd'
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.add_message(request, messages.SUCCESS, "登録が完了しました。")
        return redirect('buy:dept')
    
    context= {
        'form':form,
        'func':func,
    }
    return render(request,'buy/dept_form.html',context)

def deptdelete(request, pk):
    #urlのpkを基にdeptを取得
    dept = get_object_or_404(DeptMaster,pk=pk)

    #methot=POST、つまり送信ボタン押下時、入力内容が問題なければ
    if request.method == 'POST':
        dept.delete()
        messages.add_message(request, messages.SUCCESS, "データが正常に削除されました。")
        return redirect('buy:dept')

    context = {
        'dept':dept,
    }
    return render(request,'buy/dept_confirm_delete.html',context)

def place(request):
    context ={
        'place_list':PurchaserMaster.objects.all(),
    }
    return render(request,'buy/place.html',context)

def placeadd(request):
    form = PlaceCreateForm(request.POST or None)
    func = 'placeadd'
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.add_message(request, messages.SUCCESS, "登録が完了しました。")
        return redirect('buy:place')

    context= {
        'form':form,
        'func':func
    }
    return render(request,'buy/place_form.html',context)

def placeup(request, pk):
    #urlのpkを基にdeptを取得
    place = get_object_or_404(PurchaserMaster,pk=pk)
    #フォームに取得したdeptを紐づけ
    form = PlaceEditForm(request.POST or None, instance=place)

    #methot=POST、つまり送信ボタン押下時、入力内容が問題なければ
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.add_message(request, messages.SUCCESS, "更新が完了しました。")
        return redirect('buy:place')

    context = {
        'form':form,
    }
    return render(request,'buy/place_form.html',context)

def placedelete(request, pk):
    #urlのpkを基にdeptを取得
    place = get_object_or_404(PurchaserMaster,pk=pk)

    if request.method == 'POST':
        place.delete()
        messages.add_message(request, messages.SUCCESS, "データが正常に削除されました。")
        return redirect('buy:place')

    context = {
        'place':place,
    }
    return render(request,'buy/place_confirm_delete.html',context)

def purchases(request):
    global query
    query = ListOfPurchases.objects.order_by('-update_date_time')
    dept_list = DeptMaster.objects.distinct().values('dept')
    store_list = PurchaserMaster.objects.distinct().values('place')
    keyword = request.GET.getlist('keyword')
    if keyword:
        exclusion_list = set([' ','　',])
        q_list = ['']
        for i in keyword:
            """ 全角半角の空文字が含まれていたら無視 """
            if i in exclusion_list:
                pass
            else:
                q_list += i
        queryset = reduce( #　ListOfPurchasesのdeptで参照しているDeptMasterのdeptフィールドを指定 
                    and_, [Q(dept__dept__icontains=keyword) | Q(item_name__icontains=keyword)
                    |Q(store__place__icontains=keyword)  for keyword in q_list]
                    )
        query = query.filter(queryset)

    context ={
        'list':query,
        'dept_list':dept_list,
        'store_list':store_list,
    }
    return render(request,'buy/purchases_list.html',context)

def purchasesdelete(request, pk):
    #urlのpkを基にdeptを取得
    purchases = get_object_or_404(ListOfPurchases,pk=pk)

    if request.method == 'POST':
        purchases.delete()
        messages.add_message(request, messages.SUCCESS, "データが正常に削除されました。")
        return redirect('buy:purchases')

    context = {
        'purchases':purchases,
    }
    return render(request,'buy/purchases_confirm_delete.html',context)

def purchasesup(request, pk):
    #urlのpkを基にdeptを取得
    purchases = get_object_or_404(ListOfPurchases,pk=pk)
    #フォームに取得したdeptを紐づけ
    form = AddListForm(request.POST or None, instance=purchases)

    #methot=POST、つまり送信ボタン押下時、入力内容が問題なければ
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.add_message(request, messages.SUCCESS, "更新が完了しました。")
        return redirect('buy:purchases')

    context = {
        'form':form,
    }
    return render(request,'buy/add.html',context)

def xlsx(request):
    # 新規ブックを作成
    wb = openpyxl.Workbook()
    ws = wb.active
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=data.xlsx'
    for row in ws["A1:I1"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="centerContinuous")
    
    
    ws.cell(1, 1).value = '購入部署'
    ws.cell(1, 2).value = '注文日'
    ws.cell(1, 3).value = '購入先'
    ws.cell(1, 4).value = '購入品'
    ws.cell(1, 5).value = '単価'
    ws.cell(1, 6).value = '数量'
    ws.cell(1, 7).value = '合計'
    ws.cell(1, 8).value = '担当者'
    ws.cell(1, 9).value = '備考'
    k = 2
    for post in query:
        #外部キーは、元フィールド.参照元フィールドでないとエラー
        ws.cell(k, 1).value = post.dept.dept
        ws.cell(k, 2).value = post.order_day
        ws.cell(k, 3).value = post.store.place
        ws.cell(k, 4).value = post.item_name
        ws.cell(k, 5).value = post.unit_prise
        ws.cell(k, 6).value = post.quantity
        ws.cell(k, 7).value = post.total_due
        ws.cell(k, 8).value = post.user_id 
        ws.cell(k, 9).value = post.remarks
        k += 1
        
    # 書き込みしたデータを一旦保存
    wb.save('static/xlsx/data.xlsx')
    inputfile = 'static/xlsx/data.xlsx'
    # 保存したxlsxデータロード　これを元にxlsxデータを加工する
    wb1 = openpyxl.load_workbook(filename=inputfile)
    ws1 = wb1.worksheets[0]

    # カラム横幅指定
    for column_cells in ws1.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        length = length * 2.5
        ws1.column_dimensions[column_cells[0].column_letter].width = length

    #　xlsxファイルを保存し返す
    
    wb1.save(response)

    return response

